import io
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.colors import HexColor, white
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak

# ══════════════════════════════════════════════════════════════
# COLORS
# ══════════════════════════════════════════════════════════════
MARINE      = '#1A4B6E'
LIGHT_MARINE= '#E6F0F9'
BG          = '#FCFBF6'
SLATE       = '#2C3E50'
MUTED       = '#7F8C8D'
SUCCESS     = '#27AE60'
DANGER      = '#E74C3C'
CHART_GREY  = '#BDC3C7'

# ══════════════════════════════════════════════════════════════
# SUBJECTS
# ══════════════════════════════════════════════════════════════
MATH   = ["Arithmetic", "Algebra", "Geometry", "Calculus", "Trigonometry"]
LOGIC  = ["Logic"]
SCI    = ["Chemistry", "Biology", "Earth Science", "Physics"]
ENG    = ["English"]
ALL_SUBJ = MATH + LOGIC + SCI + ENG

def read_tracker_bytes(file_bytes):
    """Read the Excel tracker from a byte stream and return structured data."""
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    
    data = {'students': [], 'pre': {}, 'post': {}}
    
    for sheet_key, sheet_name in [('pre', 'Pre-Test Data'), ('post', 'Post-Test Data')]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        
        for row in range(2, ws.max_row + 1):
            full_name = ws.cell(row, 4).value
            if not full_name:
                continue
                
            if full_name not in data['students']:
                data['students'].append(full_name)
            
            student_data = {'total': ws.cell(row, 5).value or 0}
            for i, subj in enumerate(ALL_SUBJ):
                col_idx = 6 + i 
                val = ws.cell(row, col_idx).value
                student_data[subj] = val if val is not None else 0
            
            data[sheet_key][full_name] = student_data
    
    wb.close()
    return data

def compute_benchmarks(data, test='post'):
    """Compute cohort averages for each subject."""
    benchmarks = {}
    for subj in ALL_SUBJ:
        scores = [data[test][s].get(subj, 0) for s in data['students'] if s in data[test]]
        benchmarks[subj] = sum(scores) / len(scores) if scores else 0
    return benchmarks

def make_chart(subjects, student_scores, benchmark_scores, title, chart_width=5.5, chart_height=None):
    """Generate a horizontal bar chart as a BytesIO PNG image."""
    n = len(subjects)
    if chart_height is None:
        chart_height = max(1.2, 0.55 * n + 0.6)
    
    fig, ax = plt.subplots(figsize=(chart_width, chart_height))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)
    
    y_pos = range(n)
    bar_height = 0.35
    
    bars1 = ax.barh([y - bar_height/2 for y in y_pos], student_scores, bar_height,
                     label='Student', color=MARINE, edgecolor='none')
    bars2 = ax.barh([y + bar_height/2 for y in y_pos], benchmark_scores, bar_height,
                     label='Benchmark', color=CHART_GREY, edgecolor='none')
    
    # Data labels
    for bar in bars1:
        w = bar.get_width()
        ax.text(w + 1, bar.get_y() + bar.get_height()/2, f'{w:.0f}',
                va='center', fontsize=7, color=SLATE, fontfamily='sans-serif')
    for bar in bars2:
        w = bar.get_width()
        ax.text(w + 1, bar.get_y() + bar.get_height()/2, f'{w:.0f}',
                va='center', fontsize=7, color=MUTED, fontfamily='sans-serif')
    
    ax.set_yticks(y_pos)
    ax.set_yticklabels(subjects, fontsize=8, fontfamily='sans-serif', color=SLATE)
    ax.invert_yaxis()
    
    ax.set_title(title, loc='left', fontsize=12, fontweight='bold', color=MARINE, fontfamily='sans-serif', pad=15)
    
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.tick_params(axis='x', which='both', bottom=False, labelbottom=False)
    ax.tick_params(axis='y', which='both', left=False)
    ax.set_xlim(0, max(max(student_scores, default=0), max(benchmark_scores, default=0)) * 1.25)
    
    ax.legend(loc='upper right', fontsize=8, frameon=False, bbox_to_anchor=(1.0, 1.15))
    
    plt.tight_layout(pad=1.0)
    
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=300, transparent=True)
    plt.close(fig)
    buf.seek(0)
    return buf

def generate_pdf_bytes(excel_bytes, student_name, report_type='both'):
    """Generate the report card PDF and return it as a byte stream."""
    data = read_tracker_bytes(excel_bytes)
    return generate_pdf_from_data(data, student_name, report_type)

def generate_pdf_from_data(data, student_name, report_type='both'):
    """Generate the report card PDF from parsed data and return it as a byte stream."""
    
    pre = data['pre'].get(student_name, {})
    post = data['post'].get(student_name, {})
    
    pre_total = pre.get('total', 0)
    post_total = post.get('total', 0)
    improvement = post_total - pre_total
    
    pdf_buffer = io.BytesIO()
    
    doc = SimpleDocTemplate(
        pdf_buffer, pagesize=letter,
        topMargin=0.4*inch, bottomMargin=0.4*inch,
        leftMargin=0.6*inch, rightMargin=0.6*inch
    )
    
    elements = []
    page_w = letter[0] - 1.2*inch  # usable width
    
    styles = getSampleStyleSheet()
    s_title = ParagraphStyle('title_s', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=18, textColor=HexColor(MARINE), alignment=TA_LEFT, spaceAfter=2)
    s_subtitle = ParagraphStyle('subtitle_s', parent=styles['Normal'], fontName='Helvetica', fontSize=11, textColor=HexColor(MUTED), alignment=TA_LEFT, spaceAfter=0)
    s_label = ParagraphStyle('label_s', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, textColor=HexColor(MUTED), alignment=TA_LEFT, spaceBefore=4, spaceAfter=1)
    s_value = ParagraphStyle('value_s', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=14, textColor=HexColor(SLATE), alignment=TA_LEFT, spaceAfter=2)
    s_metric_big = ParagraphStyle('metric_big', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=22, textColor=HexColor(MARINE), alignment=TA_CENTER, spaceBefore=2, spaceAfter=0)
    s_metric_label = ParagraphStyle('metric_lbl', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=HexColor(MUTED), alignment=TA_CENTER, spaceAfter=0)
    s_section = ParagraphStyle('section_s', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11, textColor=HexColor(MARINE), alignment=TA_LEFT, spaceBefore=12, spaceAfter=6)
    s_feedback = ParagraphStyle('feedback_s', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=HexColor(SLATE), alignment=TA_LEFT, spaceBefore=4, spaceAfter=4, leading=13)

    # Header section
    header_bar = Table([['']], colWidths=[page_w], rowHeights=[4])
    header_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor(MARINE)),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_bar)
    elements.append(Spacer(1, 10))
    
    title_text = "STUDENT PROGRESS CARD"
    if report_type == 'pre': title_text = "PRE-TEST REPORT CARD"
    elif report_type == 'post': title_text = "POST-TEST REPORT CARD"
    
    title_data = [
        [Paragraph("SHORE SKWELA 5.0", s_title),
         Paragraph(title_text, ParagraphStyle('rt', parent=s_subtitle, alignment=TA_RIGHT))]
    ]
    title_tbl = Table(title_data, colWidths=[page_w * 0.5, page_w * 0.5])
    title_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, HexColor(LIGHT_MARINE)),
    ]))
    elements.append(title_tbl)
    elements.append(Spacer(1, 12))
    
    # Student info
    info_data = [[Paragraph("STUDENT NAME", s_label)], [Paragraph(student_name, s_value)]]
    info_tbl = Table(info_data, colWidths=[page_w])
    info_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 1), (-1, 1), HexColor(LIGHT_MARINE)),
        ('TOPPADDING', (0, 1), (-1, 1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 6),
        ('LEFTPADDING', (0, 0), (-1, 0), 0),
        ('LEFTPADDING', (0, 1), (-1, 1), 8),
    ]))
    elements.append(info_tbl)
    elements.append(Spacer(1, 14))
    
    # Metrics
    imp_color = SUCCESS if improvement >= 0 else DANGER
    status_text = "IMPROVED" if improvement >= 0 else "NEEDS HELP"
    
    if report_type == 'both':
        metric_headers = ["PRE-TEST TOTAL", "POST-TEST TOTAL", "IMPROVEMENT", "STATUS"]
        metric_values = [
            Paragraph(f"{pre_total}", s_metric_big),
            Paragraph(f"{post_total}", s_metric_big),
            Paragraph(f"{'+' if improvement > 0 else ''}{improvement}", ParagraphStyle('imp', parent=s_metric_big, textColor=HexColor(imp_color))),
            Paragraph(status_text, ParagraphStyle('st', parent=s_metric_label, fontSize=11, fontName='Helvetica-Bold', textColor=HexColor(MARINE))),
        ]
        col_w = [page_w * 0.25] * 4
        bg_colors = [HexColor(LIGHT_MARINE), HexColor(LIGHT_MARINE), HexColor('#EAFAF1') if improvement >= 0 else HexColor('#FDEDEC'), HexColor(LIGHT_MARINE)]
    else:
        total = pre_total if report_type == 'pre' else post_total
        label = "PRE-TEST TOTAL" if report_type == 'pre' else "POST-TEST TOTAL"
        metric_headers = [label]
        metric_values = [Paragraph(f"{total}", s_metric_big)]
        col_w = [page_w]
        bg_colors = [HexColor(LIGHT_MARINE)]

    metric_data = [[Paragraph(h, s_metric_label) for h in metric_headers], metric_values]
    metric_tbl = Table(metric_data, colWidths=col_w)
    
    style_cmds = [
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 1), (-1, 1), 8), ('BOTTOMPADDING', (0, 1), (-1, 1), 8),
    ]
    for i, color in enumerate(bg_colors): style_cmds.append(('BACKGROUND', (i, 1), (i, 1), color))
    metric_tbl.setStyle(TableStyle(style_cmds))
    elements.append(metric_tbl)
    elements.append(Spacer(1, 16))
    
    # Grade Table
    elements.append(Paragraph("SUBJECT BREAKDOWN", s_section))
    if report_type == 'both':
        tbl_header = ["Subject", "Pre-Test", "Post-Test", "Remarks"]
        tbl_col_widths = [page_w * 0.35, page_w * 0.18, page_w * 0.18, page_w * 0.29]
    else:
        label = "Pre-Test Score" if report_type == 'pre' else "Post-Test Score"
        tbl_header = ["Subject", label]
        tbl_col_widths = [page_w * 0.6, page_w * 0.4]

    tbl_data = [tbl_header]
    for i, subj in enumerate(ALL_SUBJ):
        pre_score = pre.get(subj, 0)
        post_score = post.get(subj, 0)
        if report_type == 'both':
            if post_score > pre_score: remark = "Improved"
            elif post_score == pre_score: remark = "Maintained"
            else: remark = "Review Needed"
            tbl_data.append([subj, str(pre_score), str(post_score), remark])
        else:
            score = pre_score if report_type == 'pre' else post_score
            tbl_data.append([subj, str(score)])
    
    grade_tbl = Table(tbl_data, colWidths=tbl_col_widths)
    tbl_style = [
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('TEXTCOLOR', (0, 0), (-1, 0), white), ('BACKGROUND', (0, 0), (-1, 0), HexColor(MARINE)),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9), ('TEXTCOLOR', (0, 1), (-1, -1), HexColor(SLATE)),
        ('TOPPADDING', (0, 0), (-1, -1), 5), ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 8), ('LINEBELOW', (0, -1), (-1, -1), 1, HexColor(MARINE)),
    ]
    for i in range(1, len(tbl_data)):
        bg = HexColor(LIGHT_MARINE) if i % 2 == 0 else white
        tbl_style.append(('BACKGROUND', (0, i), (-1, i), bg))
    grade_tbl.setStyle(TableStyle(tbl_style))
    elements.append(grade_tbl)
    elements.append(Spacer(1, 16))
    
    # Charts
    elements.append(PageBreak())
    elements.append(Paragraph("MASTERY ANALYSIS", s_section))
    if report_type == 'pre': desc_text = "The charts below show the student's Pre-Test performance."
    elif report_type == 'post': desc_text = "The charts below show the student's Post-Test performance."
    else: desc_text = "The charts below compare the student's Post-Test performance against the cohort average."
    elements.append(Paragraph(desc_text, s_feedback))
    elements.append(Spacer(1, 10))
    
    w_inch, h_math_sci, h_log_eng = 7.0, 2.0, 1.0
    if report_type == 'pre': chart_scores, chart_benchmarks = pre, compute_benchmarks(data, 'pre')
    else: chart_scores, chart_benchmarks = post, compute_benchmarks(data, 'post')
    
    math_scores = [chart_scores.get(s, 0) for s in MATH]; math_bench = [chart_benchmarks.get(s, 0) for s in MATH]
    sci_scores = [chart_scores.get(s, 0) for s in SCI]; sci_bench = [chart_benchmarks.get(s, 0) for s in SCI]
    logic_scores = [chart_scores.get(s, 0) for s in LOGIC]; logic_bench = [chart_benchmarks.get(s, 0) for s in LOGIC]
    eng_scores = [chart_scores.get(s, 0) for s in ENG]; eng_bench = [chart_benchmarks.get(s, 0) for s in ENG]
    
    charts_data = [
        [Image(make_chart(MATH, math_scores, math_bench, "Mathematics", w_inch, h_math_sci), width=w_inch*inch, height=h_math_sci*inch)],
        [Image(make_chart(SCI, sci_scores, sci_bench, "Science", w_inch, h_math_sci), width=w_inch*inch, height=h_math_sci*inch)],
        [Image(make_chart(LOGIC, logic_scores, logic_bench, "Logic", w_inch, h_log_eng), width=w_inch*inch, height=h_log_eng*inch)],
        [Image(make_chart(ENG, eng_scores, eng_bench, "English", w_inch, h_log_eng), width=w_inch*inch, height=h_log_eng*inch)],
    ]
    charts_tbl = Table(charts_data, colWidths=[page_w])
    charts_tbl.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(charts_tbl)
    elements.append(Spacer(1, 14))
    
    # Feedback
    elements.append(Paragraph("TEACHER'S NOTES", s_section))
    test_scores = {s: chart_scores.get(s, 0) for s in ALL_SUBJ}
    strongest = max(test_scores, key=test_scores.get)
    weakest = min(test_scores, key=test_scores.get)
    
    fb = ""
    if report_type == 'both': fb = f"Excellent progress from Pre-Test to Post-Test! " if improvement > 10 else f"Targeted review is recommended to boost overall performance. "
    elif report_type == 'pre': fb = f"Initial assessment completed. "
    elif report_type == 'post': fb = f"Final assessment completed. "
        
    fb += f"This student shows strongest mastery in <b>{strongest}</b>. The area requiring the most attention is <b>{weakest}</b>. "
    if report_type != 'post': fb += f"Consider additional one-on-one tutoring in this subject during the upcoming review sessions."
    
    feedback_tbl = Table([[Paragraph(fb, s_feedback)]], colWidths=[page_w])
    feedback_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor(LIGHT_MARINE)),
        ('TOPPADDING', (0, 0), (-1, -1), 10), ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 12), ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(feedback_tbl)
    elements.append(Spacer(1, 8))
    
    bottom_bar = Table([['']], colWidths=[page_w], rowHeights=[4])
    bottom_bar.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), HexColor(MARINE)),
        ('LEFTPADDING', (0, 0), (-1, -1), 0), ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0), ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(bottom_bar)
    
    # Render PDF into buffer
    doc.build(elements)
    
    # Return raw bytes
    pdf_buffer.seek(0)
    return pdf_buffer.read()
